from openpyxl import load_workbook
wb2 = load_workbook('abc.xlsx')
ws=wb2.active
#print(wb2.sheetnames)
c = ws['A1']
print (c.value)

