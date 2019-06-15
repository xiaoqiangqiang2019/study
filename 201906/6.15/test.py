from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")
ws.title = "New Title"
ws.sheet_properties.tabColor = "1072BA"
#ws3 = wb["New Title"]
# print(wb.sheetnames)
#for sheet in wb:
#    print(sheet.title)
c = ws['A4']
ws['A4'] = 4
d = ws.cell(row=4, column=2, value=10)
wb.save('balances.xlsx')
